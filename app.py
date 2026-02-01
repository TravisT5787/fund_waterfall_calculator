"""
Closed-End Fund Waterfall Calculator
Tests waterfall, income allocations, management fees, and ownership percentages.
"""

import streamlit as st
import pandas as pd
import io
import json

st.set_page_config(page_title="Fund Waterfall Calculator", page_icon="📊", layout="wide")

st.title("📊 Closed-End Fund Waterfall Calculator")
st.markdown("Test waterfall calculations, income allocations, management fees, and ownership percentages")
st.markdown("---")

# =============================================================================
# SIDEBAR CONFIGURATION
# =============================================================================
st.sidebar.header("Configuration")

mgmt_fee_basis = st.sidebar.selectbox(
    "Management Fee Basis",
    options=["Committed Capital", "Contributed Capital", "Current Equity"],
    index=0,
    help="Basis for calculating management fees per investor"
)

st.sidebar.markdown("---")
st.sidebar.subheader("AI Configuration")
st.sidebar.markdown("*Required for waterfall interpretation*")

api_key = st.sidebar.text_input(
    "Anthropic API Key",
    type="password",
    help="Your Anthropic API key for Claude AI waterfall interpretation"
)

# =============================================================================
# HELPER FUNCTIONS
# =============================================================================
def parse_float(val, default=0.0):
    """Parse string to float, handling commas and currency symbols."""
    if val is None or str(val).strip() == '':
        return default
    try:
        cleaned = str(val).replace(',', '').replace('$', '').replace('%', '').strip()
        return float(cleaned)
    except:
        return default

def interpret_waterfall_terms(terms_text, api_key):
    """Send waterfall terms to Claude API for interpretation."""
    if not api_key:
        return None, "Please enter your Anthropic API key in the sidebar."

    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)

        prompt = f"""Analyze the following waterfall/carried interest terms from a fund document and extract the key parameters. Return a JSON object with these fields (use null if not specified):

- preferred_return_rate: The preferred return rate as a decimal (e.g., 0.08 for 8%)
- hurdle_rate: The hurdle rate as a decimal if different from preferred return
- gp_catchup_percentage: The GP catch-up percentage as a decimal (e.g., 1.0 for 100% catch-up)
- carried_interest_rate: The carried interest/promote rate as a decimal (e.g., 0.20 for 20%)
- return_of_capital_first: Boolean - whether capital must be returned before profit split
- clawback_provision: Boolean - whether there's a GP clawback provision
- distribution_priority: Array of strings describing the waterfall tiers in order
- notes: Any important clarifications or unusual terms

Waterfall Terms:
{terms_text}

Return ONLY valid JSON, no other text."""

        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=1024,
            messages=[
                {"role": "user", "content": prompt}
            ]
        )

        response_text = message.content[0].text

        # Try to parse JSON from response
        try:
            # Handle case where response might have markdown code blocks
            if "```json" in response_text:
                response_text = response_text.split("```json")[1].split("```")[0]
            elif "```" in response_text:
                response_text = response_text.split("```")[1].split("```")[0]

            parsed = json.loads(response_text.strip())
            return parsed, None
        except json.JSONDecodeError as e:
            return None, f"Could not parse AI response as JSON: {str(e)}\n\nRaw response: {response_text}"

    except ImportError:
        return None, "Anthropic library not installed. Run: pip install anthropic"
    except Exception as e:
        return None, f"API Error: {str(e)}"

def calculate_carried_interest(investors_df, waterfall_params, total_distributions, investor_exemptions=None):
    """Calculate carried interest based on interpreted waterfall parameters."""
    if not waterfall_params:
        return investors_df, "No waterfall parameters provided."

    if investor_exemptions is None:
        investor_exemptions = {}

    # Extract parameters
    pref_return = waterfall_params.get('preferred_return_rate') or 0
    carry_rate = waterfall_params.get('carried_interest_rate') or 0.20
    return_capital_first = waterfall_params.get('return_of_capital_first', True)

    calc_details = []
    calc_details.append(f"**Waterfall Parameters:**")
    calc_details.append(f"- Preferred Return: {pref_return*100:.1f}%")
    calc_details.append(f"- Carried Interest Rate: {carry_rate*100:.1f}%")
    calc_details.append(f"- Return of Capital First: {return_capital_first}")
    calc_details.append("")

    # Calculate carried interest for each investor
    carried_interest_list = []

    total_committed = investors_df['Committed Capital'].sum()

    for idx, row in investors_df.iterrows():
        investor = row['Investor']
        committed = row['Committed Capital']
        beginning_equity = row['Beginning Equity']
        contributions = row.get('Contributions', 0)
        ending_equity = row['Ending Equity']

        # Check if investor is exempt from carried interest
        if investor in investor_exemptions and investor_exemptions[investor].get('no_carry', False):
            carried_interest_list.append(0.0)
            calc_details.append(f"**{investor}:** EXEMPT from carried interest")
            continue

        # Calculate investor's share of profits
        total_invested = beginning_equity + contributions
        profit = ending_equity - total_invested

        # Simple carried interest calculation
        # If profit > preferred return threshold, apply carry rate to excess
        if pref_return > 0:
            pref_return_amount = total_invested * pref_return
            excess_profit = max(0, profit - pref_return_amount)
            carried_interest = excess_profit * carry_rate
        else:
            # No preferred return - carry on all profits
            carried_interest = max(0, profit) * carry_rate

        carried_interest_list.append(carried_interest)

        if carried_interest > 0:
            calc_details.append(f"**{investor}:** Profit ${profit:,.2f}, Carried Interest ${carried_interest:,.2f}")

    investors_df['Carried Interest'] = carried_interest_list

    return investors_df, "\n".join(calc_details)

# =============================================================================
# STEP 1: FUND TYPE SELECTION
# =============================================================================
st.header("Step 1: Fund Type & Investor Data")

fund_type = st.radio(
    "Select Fund Type",
    options=["New Fund (First Year)", "Existing Fund (Prior Year Data)"],
    horizontal=True
)

st.markdown("---")

if fund_type == "New Fund (First Year)":
    st.markdown("""
    **Upload Excel with 4 columns:**
    1. Investor (name/designation)
    2. Committed Capital
    3. Prior Year Ending Equity (can be 0)
    4. Capital Contributions (year 1 contributions)
    """)
    expected_cols = 4
else:
    st.markdown("""
    **Upload Excel with 3 columns:**
    1. Investor (name/designation)
    2. Committed Capital
    3. Prior Year Ending Equity
    """)
    expected_cols = 3

uploaded_file = st.file_uploader("Upload Investor Data (Excel)", type=['xlsx', 'xls'])

investors_df = None

if uploaded_file:
    try:
        df = pd.read_excel(uploaded_file)

        # Standardize column names
        df.columns = df.columns.str.strip()

        # Try to identify columns
        col_mapping = {}
        for col in df.columns:
            col_lower = col.lower()
            if 'investor' in col_lower or 'name' in col_lower or 'designation' in col_lower:
                col_mapping['Investor'] = col
            elif 'commit' in col_lower:
                col_mapping['Committed Capital'] = col
            elif 'ending' in col_lower or 'equity' in col_lower or 'prior' in col_lower or 'beginning' in col_lower:
                col_mapping['Beginning Equity'] = col
            elif 'contrib' in col_lower or 'call' in col_lower:
                col_mapping['Contributions'] = col

        # If we couldn't map, use position-based
        if len(col_mapping) < 3:
            if len(df.columns) >= 3:
                col_mapping = {
                    'Investor': df.columns[0],
                    'Committed Capital': df.columns[1],
                    'Beginning Equity': df.columns[2]
                }
                if len(df.columns) >= 4:
                    col_mapping['Contributions'] = df.columns[3]

        # Rename columns
        investors_df = df.rename(columns={v: k for k, v in col_mapping.items()})

        # Ensure required columns exist
        if 'Investor' not in investors_df.columns:
            investors_df['Investor'] = [f"Investor {i+1}" for i in range(len(investors_df))]

        if 'Committed Capital' not in investors_df.columns:
            st.error("Could not identify Committed Capital column")
            investors_df = None
        elif 'Beginning Equity' not in investors_df.columns:
            st.error("Could not identify Beginning/Prior Year Equity column")
            investors_df = None
        else:
            # Convert to numeric
            investors_df['Committed Capital'] = pd.to_numeric(investors_df['Committed Capital'], errors='coerce').fillna(0)
            investors_df['Beginning Equity'] = pd.to_numeric(investors_df['Beginning Equity'], errors='coerce').fillna(0)

            if 'Contributions' in investors_df.columns:
                investors_df['Contributions'] = pd.to_numeric(investors_df['Contributions'], errors='coerce').fillna(0)
            elif fund_type == "New Fund (First Year)":
                investors_df['Contributions'] = 0.0

            # Keep only relevant columns
            keep_cols = ['Investor', 'Committed Capital', 'Beginning Equity']
            if 'Contributions' in investors_df.columns:
                keep_cols.append('Contributions')
            investors_df = investors_df[keep_cols]

            # Calculate ownership percentage
            total_committed = investors_df['Committed Capital'].sum()
            if total_committed > 0:
                investors_df['Ownership %'] = investors_df['Committed Capital'] / total_committed
            else:
                investors_df['Ownership %'] = 0

            st.success(f"Loaded {len(investors_df)} investors")
            st.dataframe(investors_df, use_container_width=True, hide_index=True)

            # Summary metrics
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Total Committed Capital", f"${total_committed:,.2f}")
            with col2:
                st.metric("Total Beginning Equity", f"${investors_df['Beginning Equity'].sum():,.2f}")
            with col3:
                if 'Contributions' in investors_df.columns:
                    st.metric("Total Contributions (from file)", f"${investors_df['Contributions'].sum():,.2f}")

            # Investor exemptions
            st.markdown("---")
            st.subheader("Investor Fee/Carry Exemptions")
            st.markdown("*Check if an investor is exempt from management fees and/or carried interest*")

            # Column headers
            col1, col2, col3 = st.columns([3, 1.5, 1.5])
            with col1:
                st.markdown("**Investor**")
            with col2:
                st.markdown("**No Mgmt Fee**")
            with col3:
                st.markdown("**No Carry**")

            # Store exemptions in session state
            if 'investor_exemptions' not in st.session_state:
                st.session_state.investor_exemptions = {}

            for idx, row in investors_df.iterrows():
                investor_name = row['Investor']
                col1, col2, col3 = st.columns([3, 1.5, 1.5])

                with col1:
                    st.markdown(f"{investor_name}")

                with col2:
                    no_mgmt_fee = st.checkbox(
                        "No Mgmt Fee",
                        key=f"no_mgmt_fee_{idx}",
                        label_visibility="collapsed"
                    )

                with col3:
                    no_carry = st.checkbox(
                        "No Carry",
                        key=f"no_carry_{idx}",
                        label_visibility="collapsed"
                    )

                st.session_state.investor_exemptions[investor_name] = {
                    'no_mgmt_fee': no_mgmt_fee,
                    'no_carry': no_carry
                }

    except Exception as e:
        st.error(f"Error reading file: {str(e)}")

st.markdown("---")

# =============================================================================
# LATE INVESTORS / EQUALIZATION (Existing Funds Only)
# =============================================================================
late_investors = []
prior_pl_for_equalization = 0.0
has_late_investors = False

if fund_type == "Existing Fund (Prior Year Data)":
    st.header("Step 1b: Late Investors (Equalization)")

    has_late_investors = st.checkbox(
        "Fund allowed late investors?",
        help="Check if new investors joined after the fund's initial closing"
    )

    if has_late_investors:
        st.markdown("""
        **Late Investor Equalization:**
        - Late investors pay interest on their committed capital to compensate existing investors
        - Interest income is allocated to existing investors based on their prior ownership %
        - All inception-to-date P/L is then reallocated among ALL investors using new ownership %
        """)

        # Input for prior P/L (needed for equalization calculation)
        prior_pl_str = st.text_input(
            "Inception-to-Date P/L (excluding management fees)",
            value="",
            placeholder="Enter total prior P/L for equalization",
            help="Total fund P/L from inception through the end of prior year, excluding management fees. This is used to calculate the equalization adjustment."
        )
        prior_pl_for_equalization = parse_float(prior_pl_str, 0.0)

        st.markdown("---")

        # Initialize session state for late investors
        if 'num_late_investors' not in st.session_state:
            st.session_state.num_late_investors = 1

        # Add/Remove buttons
        col_add, col_remove, col_spacer = st.columns([1, 1, 4])
        with col_add:
            if st.button("➕ Add Late Investor"):
                st.session_state.num_late_investors += 1
                st.rerun()
        with col_remove:
            if st.button("➖ Remove Late Investor") and st.session_state.num_late_investors > 1:
                st.session_state.num_late_investors -= 1
                st.rerun()

        # Column headers
        col1, col2, col3, col4 = st.columns([2, 1.5, 1.5, 1])
        with col1:
            st.markdown("**Investor Name**")
        with col2:
            st.markdown("**Committed Capital**")
        with col3:
            st.markdown("**Contribution**")
        with col4:
            st.markdown("**Interest Rate %**")

        for i in range(st.session_state.num_late_investors):
            col1, col2, col3, col4 = st.columns([2, 1.5, 1.5, 1])

            with col1:
                late_name = st.text_input(
                    "Name",
                    key=f"late_name_{i}",
                    value="",
                    label_visibility="collapsed",
                    placeholder="Late Investor Name"
                )

            with col2:
                late_commit_str = st.text_input(
                    "Committed",
                    key=f"late_commit_{i}",
                    value="",
                    label_visibility="collapsed",
                    placeholder="Committed Capital"
                )
                late_commit = parse_float(late_commit_str, 0.0)

            with col3:
                late_contrib_str = st.text_input(
                    "Contribution",
                    key=f"late_contrib_{i}",
                    value="",
                    label_visibility="collapsed",
                    placeholder="Contribution Amount"
                )
                late_contrib = parse_float(late_contrib_str, 0.0)

            with col4:
                late_rate_str = st.text_input(
                    "Rate",
                    key=f"late_rate_{i}",
                    value="",
                    label_visibility="collapsed",
                    placeholder="e.g., 5"
                )
                late_rate = parse_float(late_rate_str, 0.0) / 100

            if late_name and late_commit > 0:
                interest_amount = late_commit * late_rate
                late_investors.append({
                    'Investor': late_name,
                    'Committed Capital': late_commit,
                    'Contribution': late_contrib,
                    'Interest Rate': late_rate,
                    'Interest Amount': interest_amount
                })

        if late_investors:
            st.markdown("---")
            st.subheader("Late Investor Summary")
            late_df = pd.DataFrame(late_investors)
            late_df['Interest Rate'] = late_df['Interest Rate'].apply(lambda x: f"{x*100:.2f}%")
            late_df['Interest Amount'] = late_df['Interest Amount'].apply(lambda x: f"${x:,.2f}")
            st.dataframe(late_df, use_container_width=True, hide_index=True)

            total_late_interest = sum(li['Interest Amount'] for li in late_investors)
            st.metric("Total Late Investor Interest (to existing investors)", f"${total_late_interest:,.2f}")

st.markdown("---")

# =============================================================================
# STEP 2: YEAR-TO-DATE ACTIVITY
# =============================================================================
st.header("Step 2: Year-to-Date Activity")

col1, col2 = st.columns(2)

with col1:
    income_str = st.text_input(
        "YTD Income (excluding management fees)",
        value="",
        placeholder="Enter total income",
        help="Total fund income for the year, before management fees"
    )
    ytd_income = parse_float(income_str, 0.0)

with col2:
    fee_rate_str = st.text_input(
        "Management Fee Rate (%)",
        value="",
        placeholder="e.g., 2 for 2%",
        help="Annual management fee rate as a percentage"
    )
    mgmt_fee_rate = parse_float(fee_rate_str, 0.0) / 100  # Convert to decimal

st.markdown("---")

# Capital calls
col1, col2 = st.columns([1, 2])
with col1:
    has_capital_calls = st.checkbox("Capital Calls Occurred?")
with col2:
    if has_capital_calls:
        capital_calls_str = st.text_input(
            "Total Capital Calls",
            value="",
            placeholder="Enter total capital called",
            key="capital_calls"
        )
        total_capital_calls = parse_float(capital_calls_str, 0.0)
    else:
        total_capital_calls = 0.0

# Distributions
col1, col2 = st.columns([1, 2])
with col1:
    has_distributions = st.checkbox("Capital Distributions Occurred?")
with col2:
    if has_distributions:
        distributions_str = st.text_input(
            "Total Distributions",
            value="",
            placeholder="Enter total distributions",
            key="distributions"
        )
        total_distributions = parse_float(distributions_str, 0.0)
    else:
        total_distributions = 0.0

st.markdown("---")

# =============================================================================
# STEP 3: WATERFALL TERMS
# =============================================================================
st.header("Step 3: Waterfall / Carried Interest Terms")

st.markdown("""
Paste the waterfall or carried interest terms from the fund's organizational documents below.
The AI will interpret the terms and extract the key parameters.
""")

# Prior ITD carried interest
prior_itd_carry_str = st.text_input(
    "Prior Year ITD Carried Interest",
    value="",
    placeholder="Enter prior inception-to-date carried interest",
    help="Cumulative carried interest from prior years (for calculating total ITD carried interest)"
)
prior_itd_carry = parse_float(prior_itd_carry_str, 0.0)

st.markdown("---")

waterfall_terms = st.text_area(
    "Waterfall Terms",
    height=200,
    placeholder="Paste the waterfall/carried interest language from the fund documents here..."
)

waterfall_params = None

if waterfall_terms:
    if st.button("🤖 Interpret Waterfall Terms", type="secondary"):
        with st.spinner("Analyzing waterfall terms..."):
            waterfall_params, error = interpret_waterfall_terms(waterfall_terms, api_key)

            if error:
                st.error(error)
            elif waterfall_params:
                st.session_state['waterfall_params'] = waterfall_params
                st.success("Waterfall terms interpreted successfully!")

# Display interpreted parameters if they exist
if 'waterfall_params' in st.session_state and st.session_state['waterfall_params']:
    waterfall_params = st.session_state['waterfall_params']

    st.subheader("Interpreted Parameters")

    col1, col2, col3 = st.columns(3)

    with col1:
        pref_return = waterfall_params.get('preferred_return_rate')
        st.metric("Preferred Return", f"{pref_return*100:.1f}%" if pref_return else "None")

    with col2:
        carry_rate = waterfall_params.get('carried_interest_rate')
        st.metric("Carried Interest Rate", f"{carry_rate*100:.1f}%" if carry_rate else "20% (default)")

    with col3:
        catchup = waterfall_params.get('gp_catchup_percentage')
        st.metric("GP Catch-up", f"{catchup*100:.0f}%" if catchup else "None")

    # Show distribution priority if available
    dist_priority = waterfall_params.get('distribution_priority')
    if dist_priority:
        st.markdown("**Distribution Priority:**")
        for i, tier in enumerate(dist_priority, 1):
            st.markdown(f"{i}. {tier}")

    # Show notes if available
    notes = waterfall_params.get('notes')
    if notes:
        st.info(f"**Notes:** {notes}")

st.markdown("---")

# =============================================================================
# STEP 4: CALCULATE
# =============================================================================
st.header("Step 4: Calculate Allocations")

if st.button("🔄 Calculate Allocations", type="primary", use_container_width=True):

    if investors_df is None:
        st.error("Please upload investor data first")
    else:
        # Create working copy
        results_df = investors_df.copy()

        # Get waterfall params from session state
        waterfall_params = st.session_state.get('waterfall_params')

        calc_log = []

        # Track equalization and late investor interest
        results_df['Late Investor Interest'] = 0.0
        results_df['Equalization Adj'] = 0.0

        # =================================================================
        # LATE INVESTOR EQUALIZATION
        # =================================================================
        if late_investors:
            calc_log.append("--- LATE INVESTOR EQUALIZATION ---")

            # Store original ownership % (before adding late investors)
            original_ownership = results_df['Ownership %'].copy()
            original_total_committed = results_df['Committed Capital'].sum()

            # Calculate total late investor interest
            total_late_interest = sum(li['Interest Amount'] for li in late_investors)
            calc_log.append(f"Total late investor interest: ${total_late_interest:,.2f}")

            # Allocate late investor interest to existing investors by their PRIOR ownership %
            results_df['Late Investor Interest'] = original_ownership * total_late_interest
            calc_log.append(f"Late interest income allocated to existing investors by prior ownership %")

            # Ensure Contributions column exists before adding late investors
            if 'Contributions' not in results_df.columns:
                results_df['Contributions'] = 0.0

            # Add late investors to the dataframe
            for li in late_investors:
                # Late investor contribution includes their interest payment
                total_contribution = li['Contribution'] + li['Interest Amount']
                new_row = {
                    'Investor': li['Investor'],
                    'Committed Capital': li['Committed Capital'],
                    'Beginning Equity': 0.0,  # Late investors have no beginning equity
                    'Ownership %': 0.0,  # Will be recalculated
                    'Contributions': total_contribution,  # Contribution + interest
                    'Late Investor Interest': -li['Interest Amount'],  # They PAY the interest (negative)
                    'Equalization Adj': 0.0,
                    'Is Late Investor': True
                }
                results_df = pd.concat([results_df, pd.DataFrame([new_row])], ignore_index=True)
                calc_log.append(f"Added late investor: {li['Investor']} (${li['Committed Capital']:,.2f} committed, ${li['Contribution']:,.2f} + ${li['Interest Amount']:,.2f} interest = ${total_contribution:,.2f} contribution)")

            # Mark existing investors and fill NaN values
            if 'Is Late Investor' not in results_df.columns:
                results_df['Is Late Investor'] = False
            results_df['Is Late Investor'] = results_df['Is Late Investor'].fillna(False)
            results_df['Late Investor Interest'] = results_df['Late Investor Interest'].fillna(0.0)
            results_df['Equalization Adj'] = results_df['Equalization Adj'].fillna(0.0)

            # Recalculate ownership % with ALL investors
            new_total_committed = results_df['Committed Capital'].sum()
            results_df['Ownership %'] = results_df['Committed Capital'] / new_total_committed
            calc_log.append(f"Recalculated ownership % with new total committed: ${new_total_committed:,.2f}")

            # Calculate equalization adjustment
            # All inception-to-date P/L reallocated using NEW ownership %
            # vs what existing investors would have received with OLD ownership %

            for idx, row in results_df.iterrows():
                if row.get('Is Late Investor', False):
                    # Late investors get their share of inception-to-date P/L
                    # based on new ownership %, this is their equalization benefit
                    equalization = row['Ownership %'] * prior_pl_for_equalization
                    results_df.at[idx, 'Equalization Adj'] = equalization
                else:
                    # Existing investors give up some of their prior P/L
                    # Difference between old allocation and new allocation
                    old_ownership = original_ownership.get(idx, 0)
                    new_ownership = row['Ownership %']
                    equalization = (new_ownership - old_ownership) * prior_pl_for_equalization
                    results_df.at[idx, 'Equalization Adj'] = equalization

            calc_log.append(f"Equalization adjustments calculated based on ${prior_pl_for_equalization:,.2f} inception-to-date P/L")
            calc_log.append("--- END EQUALIZATION ---")
        else:
            results_df['Is Late Investor'] = False

        # =================================================================
        # ALLOCATE CAPITAL CALLS
        # =================================================================
        # Ensure Contributions column exists
        if 'Contributions' not in results_df.columns:
            results_df['Contributions'] = 0.0

        if has_capital_calls and total_capital_calls > 0:
            # Only allocate to non-late investors (late investors already have contributions set)
            if late_investors:
                # Allocate capital calls only to existing investors by their ownership
                existing_mask = ~results_df['Is Late Investor']
                existing_ownership_sum = results_df.loc[existing_mask, 'Ownership %'].sum()
                if existing_ownership_sum > 0:
                    # Add to existing contributions (don't overwrite)
                    results_df.loc[existing_mask, 'Contributions'] = (
                        results_df.loc[existing_mask, 'Contributions'].fillna(0) +
                        results_df.loc[existing_mask, 'Ownership %'] / existing_ownership_sum * total_capital_calls
                    )
            else:
                results_df['Contributions'] = results_df['Contributions'].fillna(0) + results_df['Ownership %'] * total_capital_calls
            calc_log.append(f"Capital Calls: ${total_capital_calls:,.2f} allocated by ownership %")

        # Fill any remaining NaN in Contributions
        results_df['Contributions'] = results_df['Contributions'].fillna(0.0)

        # =================================================================
        # ALLOCATE DISTRIBUTIONS
        # =================================================================
        if has_distributions and total_distributions > 0:
            results_df['Distributions'] = results_df['Ownership %'] * total_distributions
            calc_log.append(f"Distributions: ${total_distributions:,.2f} allocated by ownership %")
        else:
            results_df['Distributions'] = 0.0

        # =================================================================
        # CALCULATE MANAGEMENT FEES
        # =================================================================
        # Get exemptions from session state
        investor_exemptions = st.session_state.get('investor_exemptions', {})

        if mgmt_fee_rate > 0:
            if mgmt_fee_basis == "Committed Capital":
                results_df['Mgmt Fees'] = results_df['Committed Capital'] * mgmt_fee_rate
                calc_log.append(f"Management Fees: {mgmt_fee_rate*100:.2f}% of Committed Capital")
            elif mgmt_fee_basis == "Contributed Capital":
                # Use beginning equity + contributions as contributed capital
                contributed = results_df['Beginning Equity'] + results_df['Contributions']
                results_df['Mgmt Fees'] = contributed * mgmt_fee_rate
                calc_log.append(f"Management Fees: {mgmt_fee_rate*100:.2f}% of Contributed Capital")
            else:  # Current Equity - will be calculated after P/L
                results_df['Mgmt Fees'] = 0.0  # Placeholder
                calc_log.append(f"Management Fees: {mgmt_fee_rate*100:.2f}% of Current Equity (calculated after P/L)")

            # Apply management fee exemptions
            for idx, row in results_df.iterrows():
                investor_name = row['Investor']
                if investor_name in investor_exemptions:
                    if investor_exemptions[investor_name].get('no_mgmt_fee', False):
                        results_df.at[idx, 'Mgmt Fees'] = 0.0
                        calc_log.append(f"  - {investor_name}: EXEMPT from management fees")
        else:
            results_df['Mgmt Fees'] = 0.0

        # =================================================================
        # ALLOCATE P/L (gross income before management fees)
        # =================================================================
        results_df['P/L'] = results_df['Ownership %'] * ytd_income
        calc_log.append(f"P/L: ${ytd_income:,.2f} income allocated by ownership %")

        # =================================================================
        # CALCULATE TOTAL P/L AND ENDING EQUITY
        # =================================================================
        # Total P/L = gross P/L + late investor interest + equalization - management fees
        results_df['Total P/L'] = (
            results_df['P/L']
            + results_df['Late Investor Interest']
            + results_df['Equalization Adj']
            - results_df['Mgmt Fees']
        )
        total_mgmt_fees = results_df['Mgmt Fees'].sum()
        calc_log.append(f"Total P/L: P/L + Late Interest + Equalization - Mgmt Fees (${total_mgmt_fees:,.2f} total fees)")

        results_df['Ending Equity'] = (
            results_df['Beginning Equity']
            + results_df['Contributions']
            - results_df['Distributions']
            + results_df['Total P/L']
        )

        # If mgmt fee basis is Current Equity, recalculate
        if mgmt_fee_basis == "Current Equity" and mgmt_fee_rate > 0:
            # Use ending equity as the basis (iterative approach - one pass)
            results_df['Mgmt Fees'] = results_df['Ending Equity'] * mgmt_fee_rate
            results_df['Total P/L'] = (
                results_df['P/L']
                + results_df['Late Investor Interest']
                + results_df['Equalization Adj']
                - results_df['Mgmt Fees']
            )
            results_df['Ending Equity'] = (
                results_df['Beginning Equity']
                + results_df['Contributions']
                - results_df['Distributions']
                + results_df['Total P/L']
            )

        # =================================================================
        # CALCULATE CARRIED INTEREST
        # =================================================================
        if waterfall_params:
            results_df, carry_details = calculate_carried_interest(
                results_df,
                waterfall_params,
                total_distributions,
                investor_exemptions
            )
        else:
            results_df['Carried Interest'] = 0.0
            carry_details = "No waterfall terms provided - carried interest set to $0"

        # Apply carried interest exemptions
        for idx, row in results_df.iterrows():
            investor_name = row['Investor']
            if investor_name in investor_exemptions:
                if investor_exemptions[investor_name].get('no_carry', False):
                    results_df.at[idx, 'Carried Interest'] = 0.0

        # Calculate current year and cumulative ITD carried interest
        current_year_carry = results_df['Carried Interest'].sum()
        cumulative_itd_carry = prior_itd_carry + current_year_carry

        # =================================================================
        # PREPARE OUTPUT
        # =================================================================
        # Reorder columns for display
        # Build output columns - include equalization columns if late investors exist
        output_cols = [
            'Investor',
            'Committed Capital',
            'Beginning Equity',
            'Contributions',
            'Distributions',
            'P/L',
        ]

        # Add late investor columns if there are late investors
        if late_investors:
            output_cols.extend(['Late Investor Interest', 'Equalization Adj'])

        output_cols.extend([
            'Mgmt Fees',
            'Total P/L',
            'Ending Equity',
            'Carried Interest'
        ])

        # Ensure all columns exist
        for col in output_cols:
            if col not in results_df.columns:
                results_df[col] = 0.0

        output_df = results_df[output_cols].copy()

        # Add totals row
        totals = {
            'Investor': 'TOTAL',
            'Committed Capital': output_df['Committed Capital'].sum(),
            'Beginning Equity': output_df['Beginning Equity'].sum(),
            'Contributions': output_df['Contributions'].sum(),
            'Distributions': output_df['Distributions'].sum(),
            'P/L': output_df['P/L'].sum(),
        }

        # Add late investor totals if applicable
        if late_investors:
            totals['Late Investor Interest'] = output_df['Late Investor Interest'].sum()
            totals['Equalization Adj'] = output_df['Equalization Adj'].sum()

        totals.update({
            'Mgmt Fees': output_df['Mgmt Fees'].sum(),
            'Total P/L': output_df['Total P/L'].sum(),
            'Ending Equity': output_df['Ending Equity'].sum(),
            'Carried Interest': output_df['Carried Interest'].sum()
        })
        output_df = pd.concat([output_df, pd.DataFrame([totals])], ignore_index=True)

        # Store for export
        st.session_state['output_df'] = output_df
        st.session_state['calc_log'] = calc_log
        st.session_state['carry_details'] = carry_details

        # =================================================================
        # DISPLAY RESULTS
        # =================================================================
        st.markdown("---")
        st.header("📈 Allocation Results")

        # Format for display
        display_df = output_df.copy()
        for col in display_df.columns:
            if col != 'Investor':
                display_df[col] = display_df[col].apply(
                    lambda x: f"${x:,.2f}" if isinstance(x, (int, float)) else x
                )

        st.dataframe(display_df, use_container_width=True, hide_index=True)

        # Summary metrics
        st.markdown("---")
        col1, col2, col3, col4 = st.columns(4)

        with col1:
            st.metric("Total Beginning Equity", f"${totals['Beginning Equity']:,.2f}")
        with col2:
            st.metric("Net Contributions", f"${totals['Contributions'] - totals['Distributions']:,.2f}")
        with col3:
            st.metric("Total P/L (net of fees)", f"${totals['Total P/L']:,.2f}")
        with col4:
            st.metric("Total Ending Equity", f"${totals['Ending Equity']:,.2f}")

        # Carried interest details
        st.markdown("---")
        st.subheader("Carried Interest Calculation")
        st.markdown(carry_details)

        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Current Year Carried Interest", f"${current_year_carry:,.2f}")
        with col2:
            st.metric("Prior ITD Carried Interest", f"${prior_itd_carry:,.2f}")
        with col3:
            st.metric("Cumulative ITD Carried Interest", f"${cumulative_itd_carry:,.2f}")

        # Calculation log
        st.markdown("---")
        with st.expander("View Calculation Details"):
            for log_entry in calc_log:
                st.markdown(f"- {log_entry}")

# =============================================================================
# EXPORT
# =============================================================================
if 'output_df' in st.session_state:
    st.markdown("---")
    st.subheader("📥 Download Results")

    output_buffer = io.BytesIO()
    with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
        # Get the output dataframe and build a new one with PBC/Variance columns
        output_df = st.session_state['output_df'].copy()

        # Build new column order with PBC Input and Variance columns
        new_columns = []
        pbc_columns = []  # Track which columns are PBC Input
        variance_columns = []  # Track which columns are Variance
        calc_columns = {}  # Map variance column to its calculation column

        for col in output_df.columns:
            new_columns.append(col)
            # Add PBC Input and Variance after these specific columns
            if col in ['Mgmt Fees', 'Ending Equity', 'Carried Interest']:
                pbc_col_name = f'{col} PBC Input'
                var_col_name = f'{col} Variance'
                new_columns.append(pbc_col_name)
                new_columns.append(var_col_name)
                pbc_columns.append(pbc_col_name)
                variance_columns.append(var_col_name)
                calc_columns[var_col_name] = col

        # Create the expanded dataframe
        excel_df = pd.DataFrame()
        for col in new_columns:
            if col in output_df.columns:
                excel_df[col] = output_df[col]
            else:
                excel_df[col] = None  # Blank columns for PBC Input and Variance

        excel_df.to_excel(writer, index=False, sheet_name='Allocations')

        # Format the Excel sheet
        workbook = writer.book
        worksheet = writer.sheets['Allocations']

        from openpyxl.styles import PatternFill, Font
        from openpyxl.utils import get_column_letter

        # Define styles
        pbc_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')  # Light blue/gray
        red_font = Font(color='FF0000')
        header_font = Font(bold=True)
        currency_format = '$#,##0.00'

        num_rows = len(excel_df) + 1  # +1 for header

        # Get column indices (1-based for Excel)
        col_indices = {col: idx + 1 for idx, col in enumerate(new_columns)}

        # Format PBC Input columns
        for pbc_col in pbc_columns:
            col_idx = col_indices[pbc_col]
            # Header
            header_cell = worksheet.cell(row=1, column=col_idx)
            header_cell.value = 'PBC Input'
            header_cell.font = header_font
            header_cell.fill = pbc_fill
            # Data cells
            for row in range(2, num_rows + 1):
                cell = worksheet.cell(row=row, column=col_idx)
                cell.fill = pbc_fill
                cell.number_format = currency_format
            worksheet.column_dimensions[get_column_letter(col_idx)].width = 14

        # Format Variance columns with formulas
        for var_col in variance_columns:
            var_col_idx = col_indices[var_col]
            calc_col = calc_columns[var_col]
            calc_col_idx = col_indices[calc_col]
            pbc_col_idx = var_col_idx - 1  # PBC Input is always right before Variance

            calc_col_letter = get_column_letter(calc_col_idx)
            pbc_col_letter = get_column_letter(pbc_col_idx)

            # Header
            header_cell = worksheet.cell(row=1, column=var_col_idx)
            header_cell.value = 'Variance'
            header_cell.font = header_font

            # Data cells with formula
            for row in range(2, num_rows + 1):
                cell = worksheet.cell(row=row, column=var_col_idx)
                # Formula: if PBC is blank, show blank; otherwise calc - PBC
                cell.value = f'=IF({pbc_col_letter}{row}="","",{calc_col_letter}{row}-{pbc_col_letter}{row})'
                cell.font = red_font
                cell.number_format = currency_format
            worksheet.column_dimensions[get_column_letter(var_col_idx)].width = 14

        # Format all other numeric columns with currency
        for col in new_columns:
            if col not in pbc_columns and col not in variance_columns and col != 'Investor':
                col_idx = col_indices[col]
                for row in range(2, num_rows + 1):
                    cell = worksheet.cell(row=row, column=col_idx)
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        cell.number_format = currency_format

        # Auto-fit columns
        for col_idx in range(1, len(new_columns) + 1):
            col_letter = get_column_letter(col_idx)
            max_length = 0
            for row in range(1, num_rows + 1):
                cell = worksheet.cell(row=row, column=col_idx)
                try:
                    cell_len = len(str(cell.value)) if cell.value else 0
                    if cell_len > max_length:
                        max_length = cell_len
                except:
                    pass
            if max_length > 0 and col_letter not in [get_column_letter(col_indices[c]) for c in pbc_columns + variance_columns]:
                worksheet.column_dimensions[col_letter].width = max(max_length + 2, 12)

        # Calculation details sheet
        calc_log = st.session_state.get('calc_log', [])
        calc_df = pd.DataFrame({'Calculation Step': calc_log})
        calc_df.to_excel(writer, index=False, sheet_name='Calculation Details')

    output_buffer.seek(0)

    st.download_button(
        label="📥 Download Excel Results",
        data=output_buffer,
        file_name="fund_waterfall_results.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# Footer
st.markdown("---")
st.markdown("""
<div style='text-align: center; color: gray;'>
    <small>Closed-End Fund Waterfall Calculator | AI-Powered Waterfall Interpretation</small>
</div>
""", unsafe_allow_html=True)
